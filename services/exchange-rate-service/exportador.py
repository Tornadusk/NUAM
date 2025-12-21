"""
Módulo de exportación para tipos de cambio.
Genera archivos PDF, Excel y HTML con los datos de tipos de cambio.
"""
from datetime import datetime
from typing import List
from io import BytesIO
from decimal import Decimal

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from jinja2 import Template


def generar_pdf(tipos_cambio: List[dict], titulo: str = "Tipos de Cambio") -> bytes:
    """
    Genera un archivo PDF con los tipos de cambio.
    
    Args:
        tipos_cambio: Lista de diccionarios con los datos de tipos de cambio
        titulo: Título del documento
    
    Returns:
        bytes: Contenido del archivo PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#FF3333'),
        spaceAfter=30,
        alignment=1,  # Centrado
    )
    
    # Contenido
    elements = []
    
    # Título
    elements.append(Paragraph(titulo, title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del documento
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    elements.append(Paragraph(f"<b>Generado el:</b> {fecha_generacion}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de datos
    if tipos_cambio:
        # Detectar formato: si tiene 'Par de Monedas' usa formato Django, si no usa formato estándar
        primer_item = tipos_cambio[0] if tipos_cambio else {}
        if 'Par de Monedas' in primer_item:
            # Formato Django (ya viene con headers como claves)
            headers = list(primer_item.keys())
            data = [headers]
            for tc in tipos_cambio:
                data.append([str(tc.get(header, 'N/A')) for header in headers])
        else:
            # Formato estándar del microservicio
            headers = ['Par de Monedas', 'Tasa', 'Fecha', 'Fuente', 'Es Simulado']
            data = [headers]
            for tc in tipos_cambio:
                par = f"{tc.get('moneda_origen', 'N/A')}/{tc.get('moneda_destino', 'N/A')}"
                tasa = str(tc.get('tasa', 'N/A'))
                fecha = tc.get('fecha', 'N/A')
                fuente = tc.get('fuente', 'N/A')
                es_simulado = 'Sí' if tc.get('es_simulado', False) else 'No'
                data.append([par, tasa, fecha, fuente, es_simulado])
        
        # Crear tabla con ancho de columnas dinámico
        num_cols = len(headers)
        col_widths = [2*inch if i == 0 else 1.2*inch for i in range(num_cols)]
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF3333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            # Filas alternadas
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(table)
    else:
        elements.append(Paragraph("No hay datos de tipos de cambio disponibles.", styles['Normal']))
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def generar_excel(tipos_cambio: List[dict], titulo: str = "Tipos de Cambio") -> bytes:
    """
    Genera un archivo Excel (.xlsx) con los tipos de cambio.
    
    Args:
        tipos_cambio: Lista de diccionarios con los datos de tipos de cambio
        titulo: Título del documento
    
    Returns:
        bytes: Contenido del archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tipos de Cambio"
    
    # Estilos
    header_fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Detectar formato y headers
    headers = []
    if tipos_cambio:
        primer_item = tipos_cambio[0]
        if 'Par de Monedas' in primer_item:
            headers = list(primer_item.keys())
        else:
            headers = ['Par de Monedas', 'Tasa', 'Fecha', 'Fuente', 'Es Simulado']
    else:
        headers = ['Par de Monedas', 'Tasa', 'Fecha', 'Fuente', 'Es Simulado']
    
    num_cols = len(headers)
    col_letter = chr(64 + num_cols) if num_cols <= 26 else 'Z'
    
    # Título
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14, color="FF3333")
    ws.merge_cells(f'A1:{col_letter}1')
    ws['A1'].alignment = center_alignment
    
    # Fecha de generación
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws['A2'] = f"Generado el: {fecha_generacion}"
    ws.merge_cells(f'A2:{col_letter}2')
    ws['A2'].alignment = center_alignment
    
    # Encabezados
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Datos
    if tipos_cambio:
        primer_item = tipos_cambio[0]
        formato_django = 'Par de Monedas' in primer_item
        
        for row_idx, tc in enumerate(tipos_cambio, start=5):
            if formato_django:
                # Formato Django: usar las claves directamente
                for col_idx, header in enumerate(headers, start=1):
                    value = tc.get(header, 'N/A')
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                    cell.alignment = center_alignment
            else:
                # Formato estándar del microservicio
                par = f"{tc.get('moneda_origen', 'N/A')}/{tc.get('moneda_destino', 'N/A')}"
                tasa = tc.get('tasa', 'N/A')
                fecha = tc.get('fecha', 'N/A')
                fuente = tc.get('fuente', 'N/A')
                es_simulado = 'Sí' if tc.get('es_simulado', False) else 'No'
                
                ws.cell(row=row_idx, column=1, value=par).alignment = center_alignment
                ws.cell(row=row_idx, column=2, value=str(tasa)).alignment = center_alignment
                ws.cell(row=row_idx, column=3, value=str(fecha)).alignment = center_alignment
                ws.cell(row=row_idx, column=4, value=fuente).alignment = center_alignment
                ws.cell(row=row_idx, column=5, value=es_simulado).alignment = center_alignment
    else:
        ws.cell(row=5, column=1, value="No hay datos de tipos de cambio disponibles.")
        ws.merge_cells(f'A5:{col_letter}5')
        ws['A5'].alignment = center_alignment
    
    # Ajustar ancho de columnas
    for col_idx in range(1, len(headers) + 1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else 'Z'
        ws.column_dimensions[col_letter].width = 20
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generar_html(tipos_cambio: List[dict], titulo: str = "Tipos de Cambio") -> str:
    """
    Genera un archivo HTML con los tipos de cambio.
    
    Args:
        tipos_cambio: Lista de diccionarios con los datos de tipos de cambio
        titulo: Título del documento
    
    Returns:
        str: Contenido del archivo HTML
    """
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    # Preparar headers dinámicamente y convertir datos a formato uniforme
    headers = []
    formato_django = False
    datos_formateados = []
    
    if tipos_cambio:
        primer_item = tipos_cambio[0]
        if 'Par de Monedas' in primer_item:
            formato_django = True
            headers = list(primer_item.keys())
            # Los datos ya están en formato Django, usar directamente
            datos_formateados = tipos_cambio
        else:
            headers = ['Par de Monedas', 'Tasa', 'Fecha', 'Fuente', 'Es Simulado']
            # Convertir formato estándar a formato Django para consistencia
            for tc in tipos_cambio:
                datos_formateados.append({
                    'Par de Monedas': f"{tc.get('moneda_origen', 'N/A')}/{tc.get('moneda_destino', 'N/A')}",
                    'Tasa': tc.get('tasa', 'N/A'),
                    'Fecha': tc.get('fecha', 'N/A'),
                    'Fuente': tc.get('fuente', 'N/A'),
                    'Es Simulado': 'Sí' if tc.get('es_simulado', False) else 'No'
                })
    else:
        headers = ['Par de Monedas', 'Tasa', 'Fecha', 'Fuente', 'Es Simulado']
    
    html_template = Template("""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ titulo }}</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f4f4f4;
            color: #333;
        }
        .container {
            max-width: 1200px;
            margin: auto;
            background-color: #fff;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        h1 {
            color: #FF3333;
            text-align: center;
            margin-bottom: 20px;
        }
        .info {
            text-align: center;
            margin-bottom: 30px;
            color: #666;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        th, td {
            border: 1px solid #ddd;
            padding: 12px;
            text-align: center;
        }
        th {
            background-color: #FF3333;
            color: white;
            font-weight: bold;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        tr:hover {
            background-color: #f5f5f5;
        }
        .footer {
            text-align: center;
            margin-top: 40px;
            font-size: 0.9em;
            color: #777;
        }
        .no-data {
            text-align: center;
            padding: 40px;
            color: #999;
            font-style: italic;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>{{ titulo }}</h1>
        <div class="info">
            <strong>Generado el:</strong> {{ fecha_generacion }}
        </div>
        
        {% if datos_formateados %}
        <table>
            <thead>
                <tr>
                    {% for header in headers %}
                    <th>{{ header }}</th>
                    {% endfor %}
                </tr>
            </thead>
            <tbody>
                {% for tc in datos_formateados %}
                <tr>
                    {% for header in headers %}
                    <td>{{ tc.get(header, 'N/A') }}</td>
                    {% endfor %}
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% else %}
        <div class="no-data">
            No hay datos de tipos de cambio disponibles.
        </div>
        {% endif %}
        
        <div class="footer">
            Reporte generado por NUAM - {{ fecha_generacion.split('/')[2].split(' ')[0] }}
        </div>
    </div>
</body>
</html>
    """)
    
    return html_template.render(
        titulo=titulo,
        fecha_generacion=fecha_generacion,
        datos_formateados=datos_formateados,
        headers=headers
    )

