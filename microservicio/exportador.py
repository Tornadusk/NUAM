"""
Microservicio de Exportación de Gráficos y Datos
Permite exportar información de gráficos en múltiples formatos: CSV, Excel, PDF, HTML
"""
import csv
import json
from io import BytesIO, StringIO
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors


class ExportadorGraficos:
    """Clase para exportar datos de gráficos en diferentes formatos"""
    
    def __init__(self, datos, titulo="Reporte de Gráficos"):
        self.datos = datos
        self.titulo = titulo
        self.timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    
    def exportar_csv(self, nombre_archivo=None):
        """
        Exporta datos a formato CSV
        """
        if nombre_archivo is None:
            nombre_archivo = f"graficos_{self.timestamp}.csv"
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        # Si los datos son una lista de diccionarios
        if isinstance(self.datos, list) and len(self.datos) > 0:
            writer = csv.DictWriter(response, fieldnames=self.datos[0].keys())
            writer.writeheader()
            writer.writerows(self.datos)
        elif isinstance(self.datos, dict):
            # Si es un diccionario, convertirlo a formato de tabla
            writer = csv.writer(response)
            for key, value in self.datos.items():
                if isinstance(value, (list, dict)):
                    writer.writerow([key, json.dumps(value, ensure_ascii=False)])
                else:
                    writer.writerow([key, value])
        else:
            writer = csv.writer(response)
            writer.writerow(['Datos', json.dumps(self.datos, ensure_ascii=False)])
        
        return response
    
    def exportar_excel(self, nombre_archivo=None, nombre_hoja="Datos"):
        """
        Exporta datos a formato Excel (.xlsx)
        """
        if nombre_archivo is None:
            nombre_archivo = f"graficos_{self.timestamp}.xlsx"
        
        # Asegurar que el nombre del archivo tenga la extensión .xlsx
        if not nombre_archivo.endswith('.xlsx'):
            if nombre_archivo.endswith('.excel'):
                nombre_archivo = nombre_archivo.replace('.excel', '.xlsx')
            else:
                nombre_archivo = f"{nombre_archivo}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = nombre_hoja
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Si los datos son una lista de diccionarios
        if isinstance(self.datos, list) and len(self.datos) > 0:
            # Escribir encabezados
            headers = list(self.datos[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Escribir datos
            for row, item in enumerate(self.datos, 2):
                for col, header in enumerate(headers, 1):
                    value = item.get(header, '')
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False)
                    ws.cell(row=row, column=col, value=value)
        
        elif isinstance(self.datos, dict):
            # Escribir como tabla clave-valor
            row = 1
            ws.cell(row=row, column=1, value="Campo").fill = header_fill
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=2, value="Valor").fill = header_fill
            ws.cell(row=row, column=2).font = header_font
            
            for key, value in self.datos.items():
                row += 1
                ws.cell(row=row, column=1, value=str(key))
                if isinstance(value, (dict, list)):
                    ws.cell(row=row, column=2, value=json.dumps(value, ensure_ascii=False))
                else:
                    ws.cell(row=row, column=2, value=str(value))
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        response['Content-Length'] = len(output.getvalue())
        
        return response
    
    def exportar_pdf(self, nombre_archivo=None):
        """
        Exporta datos a formato PDF
        """
        if nombre_archivo is None:
            nombre_archivo = f"graficos_{self.timestamp}.pdf"
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#FF3333'),
            alignment=1,  # Center
            spaceAfter=30
        )
        
        # Título
        elements.append(Paragraph(self.titulo, title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Fecha de generación
        fecha_gen = Paragraph(f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal'])
        elements.append(fecha_gen)
        elements.append(Spacer(1, 0.3*inch))
        
        # Convertir datos a tabla
        if isinstance(self.datos, list) and len(self.datos) > 0:
            # Encabezados
            headers = list(self.datos[0].keys())
            data = [headers]
            
            # Datos
            for item in self.datos[:100]:  # Limitar a 100 filas para PDF
                row = [str(item.get(header, ''))[:50] for header in headers]  # Limitar longitud
                data.append(row)
            
            # Crear tabla
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF3333')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(table)
        
        elif isinstance(self.datos, dict):
            # Crear tabla clave-valor
            data = [['Campo', 'Valor']]
            for key, value in list(self.datos.items())[:50]:  # Limitar para PDF
                value_str = json.dumps(value, ensure_ascii=False)[:100] if isinstance(value, (dict, list)) else str(value)[:100]
                data.append([str(key), value_str])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF3333')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        return response
    
    def exportar_html(self, nombre_archivo=None):
        """
        Exporta datos a formato HTML
        """
        if nombre_archivo is None:
            nombre_archivo = f"graficos_{self.timestamp}.html"
        
        html_content = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.titulo}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #FF3333;
            text-align: center;
            border-bottom: 3px solid #FF3333;
            padding-bottom: 10px;
        }}
        .meta {{
            text-align: right;
            color: #666;
            font-size: 12px;
            margin-bottom: 20px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background-color: #FF3333;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr:hover {{
            background-color: #f1f1f1;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{self.titulo}</h1>
        <div class="meta">Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}</div>
"""
        
        # Generar tabla HTML
        if isinstance(self.datos, list) and len(self.datos) > 0:
            html_content += "<table>\n"
            # Encabezados
            headers = list(self.datos[0].keys())
            html_content += "<thead><tr>\n"
            for header in headers:
                html_content += f"    <th>{header}</th>\n"
            html_content += "</tr></thead>\n<tbody>\n"
            
            # Datos
            for item in self.datos:
                html_content += "<tr>\n"
                for header in headers:
                    value = item.get(header, '')
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False, indent=2)
                    html_content += f"    <td>{value}</td>\n"
                html_content += "</tr>\n"
            
            html_content += "</tbody>\n</table>\n"
        
        elif isinstance(self.datos, dict):
            html_content += "<table>\n<thead><tr><th>Campo</th><th>Valor</th></tr></thead>\n<tbody>\n"
            for key, value in self.datos.items():
                value_str = json.dumps(value, ensure_ascii=False, indent=2) if isinstance(value, (dict, list)) else str(value)
                html_content += f"<tr><td><strong>{key}</strong></td><td>{value_str}</td></tr>\n"
            html_content += "</tbody>\n</table>\n"
        
        html_content += """
    </div>
</body>
</html>
"""
        
        response = HttpResponse(html_content, content_type='text/html; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        return response

