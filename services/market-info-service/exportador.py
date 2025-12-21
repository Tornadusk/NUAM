"""
Módulo de exportación para información de bolsas de valores.
Genera archivos PDF, Excel y HTML con los datos de mercados.
"""
from datetime import datetime
from typing import List
from io import BytesIO

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from jinja2 import Template


def generar_pdf(datos_mercado: List[dict], titulo: str = "Información de Bolsas") -> bytes:
    """
    Genera un archivo PDF con los datos de bolsas de valores.
    
    Args:
        datos_mercado: Lista de diccionarios con los datos de mercados
        titulo: Título del documento
    
    Returns:
        bytes: Contenido del archivo PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#FF3333'),
        spaceAfter=30,
        alignment=1,  # Centrado
    )
    
    # Contenido
    elements = []
    
    # Título
    elements.append(Paragraph(titulo, title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del documento
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    elements.append(Paragraph(f"<b>Generado el:</b> {fecha_generacion}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de datos
    if datos_mercado:
        # Encabezados - usar las claves del primer elemento
        headers = list(datos_mercado[0].keys())
        
        # Preparar datos
        data = [headers]
        for mercado in datos_mercado:
            row = [str(mercado.get(header, 'N/A')) for header in headers]
            data.append(row)
        
        # Crear tabla
        col_widths = [2*inch] + [1.2*inch] * (len(headers) - 1)
        table = Table(data, colWidths=col_widths[:len(headers)])
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF3333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            # Filas alternadas
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(table)
    else:
        elements.append(Paragraph("No hay datos de mercados disponibles.", styles['Normal']))
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def generar_excel(datos_mercado: List[dict], titulo: str = "Información de Bolsas") -> bytes:
    """
    Genera un archivo Excel (.xlsx) con los datos de bolsas de valores.
    
    Args:
        datos_mercado: Lista de diccionarios con los datos de mercados
        titulo: Título del documento
    
    Returns:
        bytes: Contenido del archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bolsas de Valores"
    
    # Estilos
    header_fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Título
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14, color="FF3333")
    if datos_mercado:
        num_cols = len(datos_mercado[0].keys())
        ws.merge_cells(f'A1:{chr(64 + num_cols)}1')
    ws['A1'].alignment = center_alignment
    
    # Fecha de generación
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws['A2'] = f"Generado el: {fecha_generacion}"
    if datos_mercado:
        num_cols = len(datos_mercado[0].keys())
        ws.merge_cells(f'A2:{chr(64 + num_cols)}2')
    ws['A2'].alignment = center_alignment
    
    # Encabezados y datos
    if datos_mercado:
        headers = list(datos_mercado[0].keys())
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Datos
        for row_idx, mercado in enumerate(datos_mercado, start=5):
            for col_idx, header in enumerate(headers, start=1):
                value = mercado.get(header, 'N/A')
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.alignment = center_alignment
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 20
    else:
        ws.cell(row=5, column=1, value="No hay datos de mercados disponibles.")
        ws.merge_cells('A5:E5')
        ws['A5'].alignment = center_alignment
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generar_html(datos_mercado: List[dict], titulo: str = "Información de Bolsas") -> str:
    """
    Genera un archivo HTML con los datos de bolsas de valores.
    
    Args:
        datos_mercado: Lista de diccionarios con los datos de mercados
        titulo: Título del documento
    
    Returns:
        str: Contenido del archivo HTML
    """
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    # Preparar headers dinámicamente
    headers = []
    if datos_mercado:
        headers = list(datos_mercado[0].keys())
    
    html_template = Template("""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ titulo }}</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f4f4f4;
            color: #333;
        }
        .container {
            max-width: 1200px;
            margin: auto;
            background-color: #fff;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        h1 {
            color: #FF3333;
            text-align: center;
            margin-bottom: 20px;
        }
        .info {
            text-align: center;
            margin-bottom: 30px;
            color: #666;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        th, td {
            border: 1px solid #ddd;
            padding: 12px;
            text-align: center;
        }
        th {
            background-color: #FF3333;
            color: white;
            font-weight: bold;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        tr:hover {
            background-color: #f5f5f5;
        }
        .footer {
            text-align: center;
            margin-top: 40px;
            font-size: 0.9em;
            color: #777;
        }
        .no-data {
            text-align: center;
            padding: 40px;
            color: #999;
            font-style: italic;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>{{ titulo }}</h1>
        <div class="info">
            <strong>Generado el:</strong> {{ fecha_generacion }}
        </div>
        
        {% if datos_mercado %}
        <table>
            <thead>
                <tr>
                    {% for header in headers %}
                    <th>{{ header }}</th>
                    {% endfor %}
                </tr>
            </thead>
            <tbody>
                {% for mercado in datos_mercado %}
                <tr>
                    {% for header in headers %}
                    <td>{{ mercado.get(header, 'N/A') }}</td>
                    {% endfor %}
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% else %}
        <div class="no-data">
            No hay datos de mercados disponibles.
        </div>
        {% endif %}
        
        <div class="footer">
            Reporte generado por NUAM - {{ fecha_generacion.split('/')[2].split(' ')[0] }}
        </div>
    </div>
</body>
</html>
    """)
    
    return html_template.render(
        titulo=titulo,
        fecha_generacion=fecha_generacion,
        datos_mercado=datos_mercado,
        headers=headers
    )


